from bs4 import BeautifulSoup
from collections import defaultdict
from fpdf import FPDF
import gensim
import matplotlib.pyplot as plt
import nltk
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.stem import PorterStemmer
from nltk.stem import WordNetLemmatizer
import numpy as np
import os
import openpyxl
from openpyxl.styles import Font
import pandas as pd
from PIL import Image
import re
import requests
import seaborn as sns
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer

ps = PorterStemmer()
lemmatizer = WordNetLemmatizer()

nltk.download('punkt')
nltk.download('wordnet')
nltk.download('stopwords')

## Problem 1
## Problem 2
def question1_and_2():
    # this is used for pre-processing the articles to put into a good format for our q3 algorithm. The article has
    # all punctuation and numbers removed, then each word is tokenized, stop words are removed, then each word is
    # stemmed and lemmatized. - may need to definitely reference code used here:

    def cleaning_the_text(lines):
        stopwords = nltk.corpus.stopwords.words("english")
        preprocessing_lines = re.sub(r'[^\w\s]', '', lines.lower().strip())
        preprocessing_lines = re.sub(r'[0-9]', '', preprocessing_lines)
        nltk_tokens = nltk.word_tokenize(preprocessing_lines)

        for word in nltk_tokens:
            if word in stopwords:
                nltk_tokens.remove(word)

        for preprocessing_token in range(0, len(nltk_tokens)):
            nltk_tokens[preprocessing_token] = lemmatizer.lemmatize(ps.stem(nltk_tokens[preprocessing_token]))

        article = ''
        for token in nltk_tokens:
            article = article + str(token) + ' '

        # we add a full-stop on the end of each article so in q3 when we use sent_tokenize it will treat each article as an
        # individual token.
        article = article.strip() + '. '
        return article

    # this function is just to write a list for comparison and analysis between previous code. It highlights the top
    # 100 websites and how many times the keyword appears in each.
    def write_word_dictionary(mydict):
        with open('mydict2.txt', 'w') as f:
            for x in list(reversed(list(mydict)))[0:100]:
                f.write(("{}:{}\n".format(x, mydict[x])))

    # headers() is called when reading the website content as it looks into the <h1>,<h2>,etc tags. It makes sure the
    # website has at least one <hx> (where x is a n.o between 1 and 4) with content in before it adds the content to the
    # variable "article".
    def headers(q, soup):
        header_check = ''
        for k in soup.findAll('h' + str(q)):
            header_check = header_check + k.text

        return header_check

    # when we download the content, it is a large amount of text which is spread out. We manage to look into
    # __INITIAL_DATA__ which centralises all the websites in the page content rather than trying to find them spread
    # out across the whole website content. - obtained this command from https://thetechnotreat.com/web-scraping-with-
    # python-3-requests-and-beautifulsoup-bs4/

    # It converts the text in __INITIAL__DATA__ to a string, and then we filter the text to just obtain the website links.
    # We don't want links containing ichef, .jpg, .gif, .png and these are usually useless or images so we remove.
    # The websites ending in .shtml and .sms have to be handled differently so we put them in a different list called
    # weird_websites. Any suitable website we add to the list processsed_websites.

    def finding_recommended_websites(page_soup, weird_websites):
        processed_websites = []
        script = page_soup.find('script', text=re.compile('window\.__INITIAL_DATA__', re.I | re.M))
        script_string = ""

        for char in script:
            script_string += char

        # this just gets it to the character before the first '[' bracket that will enclose all the websites we want.
        location = script_string.index('items') + 6
        string_containing_websites = ""

        bracket_count = 0
        init_count = 0

        # we add all the text in between the starting '[' and the closing ']'. This contains all the website links we want.
        while True:
            location += 1
            string_containing_websites += script_string[location]
            if script_string[location] == '[':
                bracket_count += 1
                init_count += 1
            elif script_string[location] == ']':
                bracket_count += -1

            if (bracket_count == 0) and (init_count != 0):
                break

        # then use regex to just obtain the website links and put them in a list.
        all_websites = re.findall(r'https?://[^\s<>"]+|www\.[^\s<>"]+', string_containing_websites)


        for initial_website in all_websites:
            if re.search('\.jpg', initial_website):
                all_websites.remove(initial_website)

            elif re.search('\.gif', initial_website):
                all_websites.remove(initial_website)

            elif re.search('\.png', initial_website):
                all_websites.remove(initial_website)

        for suitable_website in all_websites:
            if re.search('\.stm', suitable_website):
                weird_websites.append(suitable_website)
            elif re.search('\.shtml', suitable_website):
                weird_websites.append(suitable_website)
            else:
                processed_websites.append(suitable_website)
        return processed_websites, weird_websites

    def downloading_website_content(website_list, index):
        # we convert each website into a str that we can use as our key in website_dict
        try:
            website_string = ''
            for char in website_list[index]:
                website_string = website_string + str(char)

            r = requests.get(website_list[index], allow_redirects=True).text
            website_soup = BeautifulSoup(r, 'html.parser')

        except ConnectionError:
            print("CONNECTION ERROR: ")
            return 0, 0, True

        return website_soup, website_string, False

    def content_excluding_p_tags(website_soup, article):
        for j in website_soup.findAll('li'):
            # if the <li> tag has a <p> tag then we don't add the content into "article".
            if j.find('p'):
                continue
            else:
                article = article + ' ' + j.text

        for q in range(1, 4):
            for k in website_soup.findAll('h' + str(q)):
                if headers(q, website_soup) != '':
                    article = article + ' ' + k.text
                else:
                    break

        figcaption = website_soup.select("figcaption")
        caption = ''

        for char in figcaption:
            caption += str(char)

        return article, caption

    def processing_bad_website_content(weird_websites, i):
        bad_website = False
        article_flag = False
        unique_method = []
        soup, website_key, error_flag = downloading_website_content(weird_websites, i)

        if error_flag:
            print("oopsie")
            return 0, 0, 0, 0

        article = ''
        stop_count = 0
        bad_count = 0

        # the reason we process these websites differently is because of how soup.findAll('p') acts on them.
        # It either:
        # 1) gets all the information in one go and then break down <p> tags recursively getting repeated
        # results. E.g. If <p> tag  1 is in a <p> tag 2 , it will first look into content in both tag 1 and
        # tag 2, then it will look into content in tag 2. Although a recursive algorithm may be better at
        # solving this I haven't managed to do one.

        # 2) gets the information in sections where it gets content of one big section, and acts like 1) on
        # it, and then does this for the next section sequentially.

        # 3) Just operates like soup.findAll('p') on normal websites and gets whole content and looks at
        # each <p> tag sequentially, without any sort of recursion happening.

        # If this is the case 1) or 3), we just get the content from our first 5 in findAll('p').
        # If this is the case 2), we work through all the lines

        # In both 1 and 3, this means the second results will be in the first, so we can check this with a list.

        for j in soup.findAll('p'):
            bad_count += 1
            x = str(j.text)

            # on the second iteration, we check to see if the second result is in the first.
            # we set bad_count limit to 5. This just gives the for loop 5 chances to get a result in
            # soup.findAll('p') other than [] which is always does.

            if bad_count < 5:
                # we set y to be []
                y = unique_method

                # if it is, we must STOP collection more data and just use the information we have in unique_method
                # we decide to stop through using a flag

                # on the first or second iteration, j.text may still equal ''. Usually by the third j.text
                # contains information so y != []
                # the first time content is added into y, .find() will return -1 and .strip() in str(y) will
                # not be true as nothing is in y.

                # on the second time, if the first 10 characters in x are found in y, this will mean
                # that we are either working with case 1 or case 3 and so we set bad_website to True.
                if y:
                    if str(y).find(x[0:10]) != -1:
                        bad_website = True
                    if x[0:10].strip() in str(y):
                        bad_website = True

            # if this is true, we set it back to False but break out of the loop as we will only
            # use the content we have found first time. If case 1, this is all the content so we are fine.
            # If case 3, we will only deal with the first "section" it has decided to split into. Although
            # this means we don't get all the text content in this case, it avoided getting extremely high
            # "keyword" counts due to the recursive nature of .findAll('p'), so is a compromise.

            if bad_website:
                break

            if j.text == '':
                continue

            # unless j.text is empty or the website is "bad" (either case 1 or 3), then we add are fine.
            else:
                stop_count += 1
                # if we have decided it is not a bad website, we no longer need to add.
                # if stop_count == 2 and we have not broken out of the loop, we no longer need to use unique method.

                # we add information into unique_method the first 5 times to check case 1 or 3 above.
                # if bad_website has not been set to True, then we have case 2.

                if stop_count < 5:
                    unique_method.append(j.text)

                article = article + j.text

                # we have to under-estimate than over as this can lead to worse results
                # most often this decreases by just ignoring the side links which I think is fair.
                if len(j.text) > 1000:
                    break

        # most of the time for these websites they don't consider 'title' in a <p> tag so we add it on.
        # I have seen this happen more times than not so believe it to increase accuracy.
        for l in soup.findAll('title'):
            article = article + ' ' + l.text

        return article, article_flag, soup, website_key


    def finding_word_count(article, caption, keyword, website_dict, necessary_words, influencing_words, article_flag,
                           website_string):
        word_count = 0
        # keyword: necessary_list is the key:value pair in the dictionary necessary_words.
        # this gets all the "necessary" words associated with the keyword
        necessary_list = necessary_words[keyword]

        # if the necessary_list is empty, we set article_flag to True. Otherwise we only set article_flag
        # to True if the website contains at least 1 word in the necessary_word list at least once.
        # otherwise we ignore that website as it is not relevant.

        if not necessary_list:
            article_flag = True

        else:
            for necessary_word in necessary_list:
                if (caption.lower().count(necessary_word) + article.lower().count(necessary_word)) > 0:
                    word_count += caption.lower().count(necessary_word) + article.lower().count(
                        necessary_word)
                    article_flag = True

        # if article_flag is true, this means we have a website that contains one necessary_word or has
        # no necessary_words to check. To then rank relevancy further, we do the same as we did for
        # "necessary_words" but instead with "influencing_words".

        # word_count will be equal to the number of times all "necessary and influencing words" appear
        # in a webpage.

        if article_flag:
            influencing_list = influencing_words[keyword]
            if not influencing_list:
                pass
            else:
                for influencing_word in influencing_list:
                    word_count += caption.lower().count(influencing_word) + article.lower().count(
                        influencing_word)

            website_dict[website_string] = word_count

        return website_dict

    def finding_word_count_for_bot(article, caption, keyword, website_dict, necessary_words, influencing_words, article_flag,
                           website_string):
        word_count = 0
        # keyword: necessary_list is the key:value pair in the dictionary necessary_words.
        # this gets all the "necessary" words associated with the keyword
        necessary_list = necessary_words[keyword]

        # if the necessary_list is empty, we set article_flag to True. Otherwise we only set article_flag
        # to True if the website contains at least 1 word in the necessary_word list at least once.
        # otherwise we ignore that website as it is not relevant.

        if not necessary_list:
            article_flag = True

        else:
            for necessary_word in necessary_list:
                if (caption.lower().count(necessary_word) + article.lower().count(necessary_word)) > 0:
                    word_count += caption.lower().count(necessary_word) + article.lower().count(
                        necessary_word)
                    article_flag = True

        # if article_flag is true, this means we have a website that contains one necessary_word or has
        # no necessary_words to check. As "bot" is not a keyword but just another way of finding extra material,
        # we add the condition that it also must contain at least one influencing word as well.

        # word_count will be equal to the number of times all "necessary and influencing words" appear
        # in a webpage.

        influence_flag = False
        if article_flag:
            influencing_list = influencing_words[keyword]
            if not influencing_list:
                pass
            else:
                for influencing_word in influencing_list:
                    if caption.lower().count(influencing_word) + article.lower().count(influencing_word) > 0:
                        word_count += caption.lower().count(influencing_word) + article.lower().count(influencing_word)
                        influence_flag = True
            if influence_flag:
                website_dict[website_string] = word_count

        return website_dict

    # this function goes through each page in the bbc searchbar for the keyword, then uses
    # finding_recommended_websites() to extract the normal webpages and the "weird" ones (.sms or .shtml).
    # We do this as this function then downloads the contents from each webpage, reads the content and counts how many
    # times the "keyword" appears in it's content. We then add a string of the website as a key, and the "keyword count"
    # in that website as a value into the dictionary website_dict.
    # The reason we divide it into the two lists is because for some .sms and .shtml files it reads the content
    # differently and outputs much higher "keyword counts" due to soup.findall('p'). We discuss this later in the code.

    def finding_webpages(keyword, necessary_words, influencing_words, word_count_function):
        website_dict = {}
        weird_websites = []
        article_flag = False
        error_flag = False

        # I DON'T THINK THIS TRY STATEMENT IS NEEDED ANYMORE.
        try:
            for number in range(1, 30):
                page = requests.get("https://www.bbc.co.uk/search?q= " + keyword + "&page=" + str(number)).text
                page_soup = BeautifulSoup(page, 'html.parser')
                processed_websites, weird_websites = finding_recommended_websites(page_soup, weird_websites)

                # if the processed_website does not contain 10 websites we will get an IndexError where we can just continue
                try:
                    for website in range(0, 10):
                        # we convert each website into a str that we can use as our key in website_dict
                        website_soup, website_string, error_flag = downloading_website_content(processed_websites,
                                                                                               website)

                        if error_flag:
                            error_flag = False
                            continue
                        # article is a string that we will add all text we find in <p>, <li>, <h1>,<h2>,<h3>,<h4>  and
                        # <figcaption> tags. We don't add any from <title> or <headers> as these are usually included in
                        # <p> tags.

                        article = ''

                        # for the ones that contain a video, we look into the synopsis-toggle-log to gage relevancy of
                        # the video.
                        # I was going to decrease relevancy of videos as I thought they were less relevant and not articles.
                        # I have decided however to ignore this.

                        # if re.search('episode-playout',r):
                        #    print("CONTAINS A VIDEO")

                        for i in website_soup.findAll('p'):
                            article = article + ' ' + i.text

                        # we use content_excluding_p_tags to get all <li>, <h1>, <h2>, <h3>, <h4> and figcaption tag
                        # information.

                        article, caption = content_excluding_p_tags(website_soup, article)

                        # the use finding_word_count_for_bot to find how many times necessary and influencing words appear in the
                        # variable article (containing website content).
                        website_dict = word_count_function(article, caption, keyword, website_dict, necessary_words,
                                                           influencing_words, article_flag, website_string)

                        # we then reset article_flag
                        article_flag = False

                except IndexError:
                    continue

            # we don't need this try and except.
            try:
                # we run through each .sms and .shtml website stored in weird_websites list.
                for i in range(0, len(weird_websites)):
                    article, article_flag, soup, website_key = processing_bad_website_content(weird_websites, i)
                    if (article == 0) and (article_flag == 0) and (soup == 0) and (website_key == 0):
                        continue

                    # we use content_excluding_p_tags to get all <li>, <h1>, <h2>, <h3>, <h4> and figcaption tag
                    # information.

                    article, caption = content_excluding_p_tags(soup, article)

                    # the use finding_word_count to find how many times necessary and influencing words appear in the
                    # variable article (containing website content).
                    website_dict = finding_word_count_for_bot(article, caption, keyword, website_dict, necessary_words,
                                                      influencing_words, article_flag, website_key)


            except IndexError:
                pass

        except IndexError:
            pass

        except ConnectionError:
            print("CONNECTION ERROR")

        except requests.exceptions.ConnectionError:
            print("The connection was aborted")

        website_dict = dict(sorted(website_dict.items(), key=lambda item: item[1]))

        return website_dict, weird_websites

    # for each of the keywords with more than one word (and thus containing white space in between), we look
    # into semantic_corpus and join the words together so they can be considered as one word.
    # Although this may affect semantics, I think for most of the terms they are used as phrases,
    # so joining them won't affect there meaning too much.

    def joining_keyword_phrases(semantic_corpus, n_grams):
        for y in n_grams:
            if y in semantic_corpus:
                semantic_corpus = semantic_corpus.replace(y, (' ' + ''.join(y.split()) + ' ').lower())

        return semantic_corpus

    def __main__():
        # decided to consider DOS attack as just denial of service. This led to better results in terms of quanitity
        # and quality but realise that it may lose me marks as not in exact format.

        keywords = ['malware', 'phishing', 'ransomware', 'spyware', 'encryption', 'denial of service',
                    'advanced persistent threat', 'malicious bot', 'targeted threat', 'computer virus']

        necessary_words = {"malware": ["malware"], "phishing": ["phishing"], "ransomware": ["ransomware"],
                           "spyware": ["spyware"], "encryption": ["encryption"], "denial of service": [],
                           "advanced persistent threat": ["threat", "advanced"], "malicious bot": ["bot"],
                           "targeted threat": ["target", "threat"], "computer virus": ["computer virus"]}

        influencing_words = {"malware": [], "phishing": [], "ransomware": [], "spyware": [], "encryption": [],
                             "denial of service": ["denial of service", "denial of service attack", " dos ",
                                                   "denial-of-service", "ddos", "(dos)"],
                             "advanced persistent threat": ["persistent"], "malicious bot": ["malicious", "infect", "harm"],
                             "targeted threat": ["cyber"], "computer virus": [" virus"]}

        bot_keyword = 'bot'
        bot_necessary_words = {"bot": ["bot"]}
        bot_influencing_words = {"bot": ["malicious", "harm"]}
        n_grams = []

        # for our algorithm in q3, it only deals with individual words rather than bigrams.
        # Instead of increasing complexity of our algorithm we instead remove the space and treat every phrase as one word.
        for x in keywords:
            if ' ' in x:
                n_grams.append(x.lower())

        # for each keyword, we build a dictionary of the top 100 articles and their word count using finding_webpages()
        # we then reverse the list, and add the top 100 word_counts together for values_total. This is because we will make
        # a list of how "relevant" the article is to the whole corpus we use in q3, by taking the
        # word_count in that article/ summation of word_counts
        # and we put each calculated value into relevancy_list. This will be used in our algorithm in q3.

        for word in keywords:
            relevancy_list = ''
            os.mkdir(str(word))
            os.chdir(str(word))
            mydict, weird_websites = finding_webpages(word, necessary_words, influencing_words, finding_word_count)
            if word == 'malicious bot':
                bot_dict, weird_websites2 = finding_webpages(bot_keyword, bot_necessary_words, bot_influencing_words,
                                                             finding_word_count_for_bot)
                # For "bot" weird_website2 = [], so we are not concerned with processing anything differently.
                # If we have already got the website in malicious_bots, we don't want it added in twice.
                # Only add the websites we need.
                for key in bot_dict.keys():
                    if key not in mydict.keys():
                        mydict.update({key: bot_dict[key]})

                # we then need to sort the dictionary again.
                mydict = dict(sorted(mydict.items(), key=lambda item: item[1]))

            values_total = 0
            for value in list(reversed(list(mydict.values())))[0:100]:
                values_total += value

            # this file is useful for seeing the top 100 articles and checking accuracy in how many times the "keyword"
            # appears in each one.
            write_word_dictionary(mydict)
            count = 0
            padding = '00'
            semantic_corpus = ''
            collection = open("collection1.txt", 'w')

            # for each of the top 100 websites in each keyword, we create a document in the format:
            # "keyword + padded number". txt e.g. malware067

            for x in list(reversed(list(mydict)))[0:100]:
                article2 = ''
                if count == 10:
                    padding = '0'
                document = open(word + padding + str(count) + '.txt', 'w')

                # need to process the articles diffently when we get the 'p' values. - FIX THIS
                try:
                    r = requests.get(x, allow_redirects=True)
                # this could be improved maybe by having a time thing. Look into this later
                except:
                    r = requests.get(x, allow_redirects=True)

                soup = BeautifulSoup(r.content, 'html.parser')
                # we then download the content of each website

                # need weird_websites_list, just stuck this in here for now
                if re.search('\.stm', x) or re.search('\.shtml', x):
                    article2, article_flag, soup, website_key = processing_bad_website_content(weird_websites,
                                                                                               weird_websites.index(x))

                else:
                    for i in soup.findAll('p'):
                        article2 = article2 + ' ' + i.text

                for q in range(1, 4):
                    for k in soup.findAll('h' + str(q)):
                        if headers(q, soup) != '':
                            article2 = article2 + ' ' + k.text
                        else:
                            break

                fig = soup.select("figcaption")
                for char in fig:
                    article2 += str(char)

                # we remove any non-ASCII characters from the text.
                encoded_string = article2.encode("ascii", "ignore")
                decode_string = encoded_string.decode()
                article2 = encoded_string.decode()

                # one of the files we will use is the semantic_corpus. This is a set of non pre-processed data that will
                # be used in our word2vec part of our algorithm in q3.
                semantic_corpus = semantic_corpus + decode_string

                # we then clean the text.
                article2 = cleaning_the_text(article2)

                # calculate relevance of the article and write into relevancy_list.
                frequency = list(reversed(list(mydict.values())))[count]
                relevancy = frequency / values_total
                formatted_relevancy = "{:.5f}".format(relevancy)
                if count == 0:
                    relevancy_list = str(formatted_relevancy)
                else:
                    relevancy_list = relevancy_list + ',' + str(formatted_relevancy)

                count += 1
                # we write the pre-processed text into our individual documents
                document.write(article2)
                collection.write(article2)

            # write a file with relevancy_list in
            with open('relevancy_list.txt', 'w') as doc:
                doc.write(relevancy_list)

            # write a file with semantic corpus in
            semantic_corpus_doc = open('semantic_corpus.txt', 'w')

            semantic_corpus = joining_keyword_phrases(semantic_corpus, n_grams)
            # we then replace all of the necessary words with the "keywords" as we assume they are semantically
            # identical.

            for z in necessary_words[word]:
                if z in semantic_corpus:
                    semantic_corpus = semantic_corpus.replace(z, word)

            semantic_corpus = joining_keyword_phrases(semantic_corpus, n_grams)

            # lastly, we replace all the influencing words with the "keywords" as we assume semantically identical.
            # Although this has limitations, such as "bot" being replaced with "malicious bot", we hope that this will
            # help the algorithm in q3.

            for z2 in influencing_words[word]:
                if z2 in semantic_corpus:
                    semantic_corpus = semantic_corpus.replace(z2, word)

            semantic_corpus = joining_keyword_phrases(semantic_corpus, n_grams)

            # we then go through each
            semantic_corpus_doc.write(str(semantic_corpus))
            os.chdir('..')

    __main__()


#question1_and_2()

## Problem 3
def question3(question4_flag):
    keywords = ['malware', 'phishing', 'ransomware', 'spyware', 'encryption', 'denial of service',
                'advanced persistent threat', 'targeted threat', 'computer virus', 'malicious bot']

    # this first function uses word2vec to vectorize each article and calculate a semantic distance between
    # the two words.
    # In keyword1 folder, we use 'semantic_corpus' as our corpus. We remove all punctuation, tokenize each sentence,
    # tokenize each word in each sentence and then use gensim.models.Word2Vec to build a Word2Vec model.

    # With our Word2Vec model, we then use gensim.similarities.similarity to calculate a semantic distance
    # between the two words. If the word is not in the corpus, then this means it does not appear in any articles.
    # In this instance we have a semantic threshold of 0.3, as a punishment score. This is variable.

    # https://www.oreilly.com/content/how-do-i-compare-document-similarity-using-python/

    similarities = []
    dict2 = defaultdict(dict)

    def word2vec_similarity(keyword1, keyword2):
        similarity_count = 0
        os.chdir(keyword1)
        f = open('semantic_corpus.txt')
        lines = str(f.readlines())
        # this only removes the punctuation instead we want to remove all but
        # words?
        # might want to remove stop words
        y = re.sub(r'[^\w\s.]', '', lines.lower().strip())
        x = sent_tokenize(y)
        words = [nltk.word_tokenize(sent) for sent in x]
        word2vec = gensim.models.Word2Vec(words, min_count=1)

        # In preprocessing semantic_corpus, for keywords that contained more than one word, we removed any whitespace.
        # We do the same for the two input words, so that we can find the semantic similarity between them.
        keyword1 = keyword1.lower().replace(' ','')
        keyword2 = keyword2.lower().replace(' ','')

        try:
            result = word2vec.wv.similarity(keyword1, keyword2)
            similarities.append([similarity_count, result])

        # semantic threshold
        except KeyError:
            result = 0.3

        dict2[keyword1][keyword2] = result

        os.chdir('..')
        return result

    # MAY NEED TO CHECK THIS FOR PLAGIARISM !!!
    # https: // dev.to / coderasha / compare - documents - similarity - using - python - nlp - 4odp
    # https: // betterprogramming.pub / introduction - to - gensim - calculating - text - similarity - 9e8b55de342d
    # this function opens the individual article, tokenizes each sentence and each word in that sentence and converts it
    # into a Bag of Words. It then checks the list, sims, which contains a comparison of each article in one directory
    # with the corpus, and chooses the max similarity. I have done this because ...
    # In the referred code, they also included a tf-idf model. However I decided not to use this because ...


    def individual_article_comparison(article, scatter_count, semantic_files_flag):
        file2_docs = []
        with open(article) as f:
            tokens = sent_tokenize(f.read())
            for line in tokens:
                file2_docs.append(line)

        for line in file2_docs:
            query_doc = [w.lower() for w in word_tokenize(line)]
            query_doc_bow = dictionary.doc2bow(query_doc)


        query_count = 0

        # FIRST PART FOR Q4 is really cool. Allows you to see scatter plot for
        # how much your article relates to each article in the corpus.
        if semantic_files_flag:
            data = []
            for x in sims[query_doc_bow]:
                query_count += 1
                data.append([query_count,x])

            frame = pd.DataFrame(data, columns= ["articles","similarity"])
            sim_scatter = sns.scatterplot(data= frame, x = "articles",y = "similarity")
            figure3 = sim_scatter.get_figure()

            if scatter_count < 10:
                figure3.savefig("00" + str(scatter_count) +".png")

            else:
                figure3.savefig("0" + str(scatter_count) +'.png')
            figure3.clear()
            return max(sims[query_doc_bow])

        else:
            return max(sims[query_doc_bow])

    # we have our corpus as our collection of articles. In pre-processing, we removed punctuation, then when we wrote an
    # article to 'collection1.txt', we added a full stop. This means when we do sent_tokenize, each token is an article.
    # We then convert these into a Bag of Words, which we then change to a Tf-Idf model using gensim.
    plot = []
    dict1 = defaultdict(dict)
    for word in keywords:
        file_docs = []
        word_result = []
        os.chdir(word)


        with open('collection1.txt') as f:
            tokens = sent_tokenize(f.read())
            for line in tokens:
                file_docs.append(line)

        gen_docs = [[w.lower() for w in word_tokenize(text)] for text in file_docs]

        dictionary = gensim.corpora.Dictionary(gen_docs)

        # this makes a csv file that has all 100 articles in their "Bag of Words" format.
        if question4_flag:
            # https://www.mygreatlearning.com/blog/bag-of-words/
            # used this to make a bag of words file in excel for visualisation
            CountVec = CountVectorizer(ngram_range=(1, 1),  # to use bigrams ngram_range=(2,2)
                                       stop_words='english')
            # transform
            Count_data = CountVec.fit_transform(tokens)

            # using panda to display the data of the bag of words we have for each corpus that we use.
            cv_dataframe = pd.DataFrame(Count_data.toarray(), columns=CountVec.get_feature_names())
            cv_dataframe.to_csv("Bag of words.xlsx")

        corpus = [dictionary.doc2bow(gen_doc) for gen_doc in gen_docs]

        # TFIDF is also a bag-of-words model but unlike the regular corpus, TFIDF down weights
        # tokens that appear frequently across documents
        # words that appear more frequently get smaller weights

        # REFERENCE:https://dev.to/coderasha/compare-documents-similarity-using-python-nlp-4odp

        dir = str(os.getcwd())
        # building the index for similarity, so we can then index each article to find out it's similarity using
        # individual article comparison.

        sims = gensim.similarities.Similarity(dir,corpus,num_features=len(dictionary))

        os.chdir('..')

        # for each comparison keyword, we get the semantic similarity by calling the function word2vec_similarity()
        # if the word = keyword we want to generate results of a scatter plot.
        for keyword in keywords:
            if word == keyword:
                semantic_files_flag = True
            else:
                semantic_files_flag = False

            similarity = word2vec_similarity(word, keyword)
            relevancy_list = ''
            os.chdir(keyword)
            summation = 0
            count = 0
            with open('relevancy_list.txt') as rl:
                relevancy_list_text = rl.read()

            relevancy_list = [value for value in relevancy_list_text.split(',')]
            relevance_check = 0

            # for each document in the comparison keyword's directory, keyword 2, we calculate how similar
            # the article is compared to keyword 1's corpus using bag of words, and multiply this by the articles
            # "relevancy" (how much it contributes depending on how many times the keyword appears in the article)

            for document in os.listdir(os.getcwd()):
                if str(keyword) in document:
                    summation += individual_article_comparison(document,count, semantic_files_flag) * np.float32(relevancy_list[count])
                    count += 1

            # we then multiply this value by how semantically similar the two words are to get our overall
            # semantic distance
            word_result.append(summation * similarity)

            dict1[word][keyword] = summation * similarity
            # this would then be multiplied by the relativity
            os.chdir('..')


    similarity_dict = pd.DataFrame(dict2)
    similarity_dict.to_excel("output_similarity.xlsx")
    # https: // pandas.pydata.org / pandas - docs / stable / user_guide / style.html

    d = pd.DataFrame(dict1)
    d.to_excel("output.xlsx")
    heatmap = sns.heatmap(d)
    fig = heatmap.get_figure()
    fig.savefig("heatmap.jpeg")
    plt.show()

## Problem 4
question4_flag = True
question3(question4_flag)

def question4():
    # this function adds the word "keyword" to the top left hand side to get it in the correct order.
    def keywords_title(filename):
        ft = Font(bold=True, color='FF0000')
        wb = openpyxl.load_workbook(filename=filename)
        ws = wb.worksheets[0]
        a1 = ws['A1']
        a1.font = ft
        ws['A1'] = 'Keywords'
        wb.save(filename=filename)

    keywords_title('output.xlsx')

    necessary_words = {"malware": 0, "phishing": 1, "ransomware": 2, "spyware": 3, "encryption": 4,"denialofservice": 5,
                       "advancedpersistentthreat": 6, "maliciousbot": 7, "targetedthreat": 8, "computervirus": 9}

    data2 = pd.read_excel("output_similarity.xlsx")
    visual_keywords = []
    visual_keywords2 = []
    for x in data2.columns:
        if x == "Unnamed: 0":
            continue
        else:
            visual_keywords.append(x)
            visual_keywords2.append(x)

    os.mkdir("files")
    os.chdir("files")
    for i in range(0, len(visual_keywords)):
        word = visual_keywords[0]
        for comp_word in visual_keywords:
            data5 = [word +" to " + comp_word, data2[word][necessary_words[comp_word]]], [comp_word +" to " + word, data2[comp_word][necessary_words[word]]]
            data6 = pd.DataFrame(data5, columns= ["keyword","similarity"])
            count = sns.barplot(data=data6, x="keyword", y="similarity")
            count.get_figure().savefig("Semantic similarity between " + word + " and " + comp_word + ".png")
            count.clear()
        visual_keywords.remove(word)

    # https://prahladyeri.com/blog/2019/10/python-recipe-combine-images-pdf.html
    # used this website to look into converting multiple images into pdf.

    pdf = FPDF()
    w, h = 0, 0
    i = 0
    for x in visual_keywords2:
        for y in visual_keywords2:
            i += 1
            fname = "Semantic similarity between " + str(x) + " and " + str(y) + ".png"
            if os.path.exists(fname):
                if i == 1:
                    cover = Image.open(fname)
                    w, h = cover.size
                    pdf = FPDF(unit="pt", format=[w, h])
                image = fname
                pdf.add_page()
                pdf.image(image, 0, 0, w, h)
            else:
                print("")

    pdf.output("output.pdf", "F")
    os.chdir('..')

    # this is a scatter graph displaying the results
    data3 = pd.read_excel("output.xlsx")
    data3.set_index('Keywords', inplace=True)
    data3 = pd.DataFrame(data3,
                         columns=["malware", "phishing", "ransomware", "spyware", "encryption", "denial of service",
                                  "advanced persistent threat", "targeted threat", "computer virus", "malicious bot"])

    x = sns.scatterplot(data=data3)
    fig = x.get_figure()
    fig.savefig("output3.png")


question4()